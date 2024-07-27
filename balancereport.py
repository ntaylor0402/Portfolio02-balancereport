import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from tkinter.filedialog import askopenfilename
import time
from columnfinder import getColumns

COLUMNS_DISPLAYED = ('Name 1', 'Posting Date', 'Amount in Local Currency', 'Reference')


def main():
    print('***** Balance Report Generator *****')

    print('Open the FBL1N spreadsheet export')
    time.sleep(.5)
    print('Press enter to open file explorer')
    input('> ')

    wb, ws = initializeWorkbook()
    columns = getColumns(ws)
    ws = deleteColumns(ws, columns)
    ws = deleteImages(ws)

    ws.delete_rows(ws.max_row)
    ws.insert_cols(4)

    ws = getBalance(ws)

    wb.save(r"C:\Users\Noah Taylor\Desktop\test.XLSX")


def initializeWorkbook():
    """prompts user to import file and returns a tuple containing a
       worksheet and workbook objects"""
    path = askopenfilename()

    wb = xl.load_workbook(path)
    ws = wb.active

    ws.delete_cols(ws.max_column)

    return wb, ws


def deleteColumns(ws, columns):
    """Deletes columns with titles not listed COLUMNS_DISPLAYED constant"""
    deleteList = []
    for key in columns.keys():
        if key not in COLUMNS_DISPLAYED:
            deleteList.append(columns[key])

    for columnIndex in sorted(deleteList, reverse=True):
        ws.delete_cols(columnIndex)
    return ws


def deleteImages(ws):
    """delete images included in the SAP export"""
    for i in range(len(ws._images)):
        del ws._images[0]
    return ws


def getBalance(ws):
    """populate a balance column with cumulative payables balance"""
    ws.cell(row=2, column=4).value = ws.cell(row=2, column=3).value
    balanceHeadingCell = ws.cell(row=1, column=4)
    balanceHeadingCell.value = 'Balance'
    balanceHeadingCell.fill = PatternFill("solid", fgColor="C0C0C0")
    balanceHeadingCell.border = Border(bottom=Side(border_style="thin", color="000000"))
    balanceHeadingCell.alignment = Alignment(horizontal='left', vertical='top')
    for i in range(3, ws.max_row + 1):
        ws.cell(row=i, column=4).value = ws.cell(row=i, column=3).value + ws.cell(row=i - 1, column=4).value
    return ws


if __name__ == '__main__':
    main()
